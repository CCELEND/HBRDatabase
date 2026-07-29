#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

import os
import configparser
import pandas as pd
import datetime
import threading
from collections import OrderedDict

from PyQt5.QtWidgets import (
    QFrame, QLabel, QPushButton, QTextEdit, QGridLayout,
    QMessageBox
)
from PyQt5.QtCore import Qt, QTimer
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from window_qt import set_window_icon, show_context_menu, creat_Toplevel
from window_qt import win_open_manage, win_close_manage, is_win_open, win_set_top
from 工具.GetEntriesGUILocal.proc import parallel_process_indexes

from 日志.advanced_logger import AdvancedLogger
logger = AdvancedLogger.get_logger(__name__)

index_wash_entries = {}
index_wash_entries_lock = threading.Lock()
index_equipments = {}
index_equipments_lock = threading.Lock()

output_text = None


def copy_text(event, text_widget):
    try:
        if isinstance(text_widget, QTextEdit):
            text_widget.copy()
    except Exception:
        pass


def paste_text(event, text_widget):
    try:
        if isinstance(text_widget, QTextEdit):
            text_widget.paste()
    except Exception:
        pass


def cut_text(event, text_widget):
    try:
        if isinstance(text_widget, QTextEdit):
            text_widget.cut()
    except Exception:
        pass


def clear_entries():
    index_wash_entries.clear()
    index_equipments.clear()


def clear_text(*text_widgets):
    for text_widget in text_widgets:
        if isinstance(text_widget, QTextEdit):
            if not text_widget.isEnabled():
                text_widget.setEnabled(True)
                text_widget.clear()
                text_widget.setEnabled(False)
            else:
                text_widget.clear()
    clear_entries()


def edit_text(text_widget, data):
    if isinstance(text_widget, QTextEdit):
        if not text_widget.isEnabled():
            text_widget.setEnabled(True)
            text_widget.clear()
            text_widget.setPlainText(str(data))
            text_widget.setEnabled(False)
        else:
            text_widget.clear()
            text_widget.setPlainText(str(data))


def print_dir(dir_data):
    data = ""
    for key, value in dir_data.items():
        data += f"{key}: {value}\n"
    edit_text(output_text, data)


def check_wash_entries(ChangeAbility_seed, ChangeAbility_index, DataCount):
    empty_variables = []
    if not ChangeAbility_seed:
        empty_variables.append('ChangeAbility_seed')
    if not ChangeAbility_index:
        empty_variables.append('ChangeAbility_index')
    if not DataCount:
        empty_variables.append('DataCount')

    empty_variables_message = "以下变量为空: " + ", ".join(empty_variables) + "\n请修改配置文件 ./工具/GetEntriesGUILocal/config.ini"
    if empty_variables:
        logger.error(str(empty_variables_message))
        QMessageBox.critical(None, "错误", empty_variables_message)
        return True
    return False


def check_equipments(RandomMainAbility_seed, RandomMainAbility_index, DataCount):
    empty_variables = []
    if not RandomMainAbility_seed:
        empty_variables.append('RandomMainAbility_seed')
    if not RandomMainAbility_index:
        empty_variables.append('RandomMainAbility_index')
    if not DataCount:
        empty_variables.append('DataCount')

    empty_variables_message = "以下变量为空: " + ", ".join(empty_variables) + "\n请修改配置文件 ./工具/GetEntriesGUILocal/config.ini"
    if empty_variables:
        logger.error(str(empty_variables_message))
        QMessageBox.critical(None, "错误", empty_variables_message)
        return True
    return False


def get_index_wash_entries():
    global index_wash_entries

    if not os.path.exists("./工具/GetEntriesGUILocal/config.ini"):
        logger.error("配置文件 ./工具/GetEntriesGUILocal/config.ini 不存在！")
        QMessageBox.critical(None, "错误", "配置文件 ./工具/GetEntriesGUILocal/config.ini 不存在！")
        return

    config = configparser.ConfigParser()
    config.read('./工具/GetEntriesGUILocal/config.ini')

    ChangeAbility_seed = config.get('ChangeAbility', 'ChangeAbility_seed', fallback="")
    ChangeAbility_index = config.get('ChangeAbility', 'ChangeAbility_index', fallback="")
    DataCount = config.get('Count', 'DataCount', fallback="")

    if check_wash_entries(ChangeAbility_seed, ChangeAbility_index, DataCount):
        return

    clear_entries()

    seed = int(ChangeAbility_seed)
    start_index = int(ChangeAbility_index)
    end_index = start_index + int(DataCount)

    starttime = datetime.datetime.now()
    temp_index_wash_entries = parallel_process_indexes(
        fun=0,
        seed=seed,
        start_index=start_index,
        end_index=end_index,
        chunk_size=10,
        max_workers=max(4, os.cpu_count())
    )
    index_wash_entries = OrderedDict(sorted(
        temp_index_wash_entries.items(),
        key=lambda x: int(x[0])
    ))
    endtime = datetime.datetime.now()

    print("use times {0:.2f}s".format((endtime - starttime).total_seconds()))
    print_dir(index_wash_entries)


def get_index_equipments():
    global index_equipments

    if not os.path.exists("./工具/GetEntriesGUILocal/config.ini"):
        logger.error("配置文件 ./工具/GetEntriesGUILocal/config.ini 不存在！")
        QMessageBox.critical(None, "错误", "配置文件 ./工具/GetEntriesGUILocal/config.ini 不存在！")
        return

    config = configparser.ConfigParser()
    config.read('./工具/GetEntriesGUILocal/config.ini')

    RandomMainAbility_seed = config.get('RandomMainAbility', 'RandomMainAbility_seed', fallback="")
    RandomMainAbility_index = config.get('RandomMainAbility', 'RandomMainAbility_index', fallback="")
    DataCount = config.get('Count', 'DataCount', fallback="")

    if check_equipments(RandomMainAbility_seed, RandomMainAbility_index, DataCount):
        return

    clear_entries()

    seed = int(RandomMainAbility_seed)
    start_index = int(RandomMainAbility_index)
    end_index = start_index + int(DataCount)

    starttime = datetime.datetime.now()
    temp_index_equipments = parallel_process_indexes(
        fun=1,
        seed=seed,
        start_index=start_index,
        end_index=end_index,
        chunk_size=10,
        max_workers=max(4, os.cpu_count())
    )
    index_equipments = OrderedDict(sorted(
        temp_index_equipments.items(),
        key=lambda x: int(x[0])
    ))
    endtime = datetime.datetime.now()

    print("use times {0:.2f}s".format((endtime - starttime).total_seconds()))
    print_dir(index_equipments)


def save_to_file():
    if index_wash_entries:
        save_index_wash_entries_to_file(index_wash_entries)
    elif index_equipments:
        save_index_equipments_to_file(index_equipments)
    else:
        QMessageBox.information(None, "提示", "无数据，请获取词条")


def fill_index_equipments_color(df, worksheet):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for idx, row in df.iterrows():
        if row['DP'] == "DP+1200":
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=idx + 2, column=col).fill = yellow_fill
        else:
            if row['第一词条'] == "第一词条+15%":
                worksheet.cell(row=idx + 2, column=2).fill = blue_fill
            if row['通常攻击攻击力'] == "通常攻击攻击力+200％":
                worksheet.cell(row=idx + 2, column=5).fill = blue_fill

    return worksheet


def save_index_equipments_to_file(index_equipments):
    df = pd.DataFrame.from_dict(index_equipments, orient='index',
                                columns=[
                                    '第一词条', 'DP', '智慧',
                                    '通常攻击攻击力', '攻击属性', '体力', '精神',
                                    '属性', '职能类型初始SP',
                                    '吊饰',
                                    '真实随机值'
                                ]
                                )
    df.reset_index(inplace=True)
    df.rename(columns={'index': '索引'}, inplace=True)

    excel_file_path = './工具/GetEntriesGUILocal/index_equipments.xlsx'
    try:
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            worksheet = fill_index_equipments_color(df, worksheet)

            column_widths = [10, 14, 12, 12, 22, 16, 12, 12, 12, 18, 12, 14]
            for i, width in enumerate(column_widths, start=1):
                worksheet.column_dimensions[get_column_letter(i)].width = width

        QMessageBox.information(None, "提示", "装备词条数据已保存至: \n./工具/GetEntriesGUILocal/index_equipments.xlsx")
    except Exception as e:
        logger.error(f"{e}\n请关闭打开的 index_equipments.xlsx 并重试")
        QMessageBox.critical(None, "错误", f"{e}\n请关闭打开的 index_equipments.xlsx 并重试")


def fill_index_wash_entries_color(df, worksheet):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for idx, row in df.iterrows():
        if '+3' in str(row['词条']) and ('+30' not in str(row['词条'])):
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=idx + 2, column=col).fill = yellow_fill

    return worksheet


def save_index_wash_entries_to_file(index_wash_entries):
    df = pd.DataFrame.from_dict(index_wash_entries, orient='index', columns=['词条', '真实随机值'])
    df.reset_index(inplace=True)
    df.rename(columns={'index': '索引'}, inplace=True)

    excel_file_path = './工具/GetEntriesGUILocal/index_wash_entries.xlsx'
    try:
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            worksheet = fill_index_wash_entries_color(df, worksheet)

            column_widths = [10, 14, 14]
            for i, width in enumerate(column_widths, start=1):
                worksheet.column_dimensions[get_column_letter(i)].width = width

        QMessageBox.information(None, "提示", "洗孔词条数据已保存至: \n./工具/GetEntriesGUILocal/index_wash_entries.xlsx")
    except Exception as e:
        logger.error(f"{e}\n请关闭打开的 index_wash_entries.xlsx 并重试")
        QMessageBox.critical(None, "错误", f"{e}\n请关闭打开的 index_wash_entries.xlsx 并重试")


def ini_config():
    if not os.path.exists("./工具/GetEntriesGUILocal/config.ini"):
        config = configparser.ConfigParser()
        config['ChangeAbility'] = {
            'ChangeAbility_seed': '',
            'ChangeAbility_index': ''
        }
        config['RandomMainAbility'] = {
            'RandomMainAbility_seed': '',
            'RandomMainAbility_index': ''
        }
        config['Count'] = {
            'DataCount': '300'
        }
        with open('./工具/GetEntriesGUILocal/config.ini', 'w') as configfile:
            config.write(configfile)


def creat_ct_win():
    global output_text

    if is_win_open('词条获取', __name__):
        win_set_top('词条获取', __name__)
        return "break"

    ct_win_frame = creat_Toplevel("词条获取", 700, 405)
    set_window_icon(ct_win_frame, "./工具/GetEntriesGUILocal/entries.ico")

    ini_config()

    layout = ct_win_frame.grid_layout
    layout.setRowStretch(0, 1)
    for col in range(4):
        layout.setColumnStretch(col, 1)

    get_entries_frame = QFrame(ct_win_frame.centralWidget())
    get_entries_frame.setLayout(QGridLayout())
    get_entries_frame.layout().setContentsMargins(0, 10, 0, 20)
    get_entries_frame.layout().setRowStretch(0, 1)
    for col in range(4):
        get_entries_frame.layout().setColumnStretch(col, 1)
    layout.addWidget(get_entries_frame, 0, 0, 1, 4)

    output_text = QTextEdit()
    output_text.setAcceptRichText(False)
    output_text.setContextMenuPolicy(Qt.CustomContextMenu)
    output_text.customContextMenuRequested.connect(
        lambda pos, tw=output_text: show_context_menu(pos, tw)
    )
    get_entries_frame.layout().addWidget(output_text, 0, 0, 1, 3)

    buttons_frame = QFrame(get_entries_frame)
    buttons_frame.setLayout(QGridLayout())
    for row in range(4):
        buttons_frame.layout().setRowStretch(row, 1)
    buttons_frame.layout().setColumnStretch(0, 1)
    get_entries_frame.layout().addWidget(buttons_frame, 0, 3)

    def run_in_thread(func):
        get_wash_entries_button.setEnabled(False)
        get_index_equipments_button.setEnabled(False)

        def wrapper():
            try:
                func()
            finally:
                QTimer.singleShot(0, lambda: get_wash_entries_button.setEnabled(True))
                QTimer.singleShot(0, lambda: get_index_equipments_button.setEnabled(True))

        thread = threading.Thread(target=wrapper, daemon=True)
        thread.start()

    get_wash_entries_button = QPushButton("获取洗孔词条")
    get_wash_entries_button.clicked.connect(lambda: run_in_thread(get_index_wash_entries))
    buttons_frame.layout().addWidget(get_wash_entries_button, 0, 0)

    get_index_equipments_button = QPushButton("获取装备词条")
    get_index_equipments_button.clicked.connect(lambda: run_in_thread(get_index_equipments))
    buttons_frame.layout().addWidget(get_index_equipments_button, 1, 0)

    clear_button = QPushButton("清空")
    clear_button.clicked.connect(lambda: clear_text(output_text))
    buttons_frame.layout().addWidget(clear_button, 2, 0)

    save_file_button = QPushButton("保存为 Excel 文件")
    save_file_button.clicked.connect(save_to_file)
    buttons_frame.layout().addWidget(save_file_button, 3, 0)

    win_open_manage(ct_win_frame, __name__)
    ct_win_frame.closeEvent = lambda ev: win_close_manage(ct_win_frame, __name__)

    return "break"

